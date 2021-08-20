import sqlite3
import openpyxl

connect = sqlite3.connect('Covidstatsdb.sqlite')
cur = connect.cursor()

#prompt user to select the state
print('NewYork = 1 \nCalifornia = 2 \nFlorida = 3 \nTexas = 4 \nSouth Dakota = 5 \nPlease select an option: ')
select = int(input())
state = ''

if select == 1:
    state = 'NYC'
elif select == 2:
    state = 'CA'
elif select == 3:
    state = 'FL'
elif select == 4:
    state = 'TX'
elif select == 5:
    state = 'SD'
else:
    print('Error did not enter a state')



cur.execute('''select Stats.submission_date, Stats.new_cases, States.name,
States.population from Stats join States on Stats.state_id = States.id WHERE state_id = ?''', (select, ))
state_info = cur.fetchall()
#print(state_info)
#returns a list of tuples with order (submission_date, new_cases, name, population)

try:
    wb = openpyxl.Workbook()
    wb.save(filename = state + '-covid-per-capita.xlsx')
except:
    wb = openpyxl.load_workbook(state + '-covid-per-capita.xlsx')


capitaD = list()
capitaBW = dict()
#calculate the daily percapita
for day in state_info:
    percapita = (day[1] / float(day[3])) * 100000.0
    capitaD.append(percapita)

#calculate the biweekly percapita
week_counter = 1
sum = 0
for day in state_info:
    sum = sum + day[1]
    #14th day
    if (week_counter % 14) == 0:
        average = ((sum / 14) / float(day[3])) * 100000.0
        capitaBW[day[0]] = capitaBW.get(day[0], average)
        week_counter = week_counter + 1
        sum = 0
    #if the counter is not divisible by 14 and is not at the end of the list
    elif (week_counter % 14) != 0 and len(state_info) > week_counter:
        week_counter = week_counter + 1
        continue
    #what to do when it arrives at the end of the list
    else:
        average = (sum / float(week_counter % 14)) / float(day[3]) * 100000.0
        capitaBW[day[0]] = capitaBW.get(day[0], average)


#insert the information into excel

ws = wb.create_sheet()
ws.title = state + '-Analysis'
extra = wb['Sheet']
wb.remove(extra)

#daily percapita
rcounter = 2
capitaIndex = 0
ws.cell(row = 1, column = 1).value = 'submission_date'
ws.cell(row = 1, column = 2).value = 'daily-percapita (per 100,000)'

for day in state_info:
    for coln in range(1,3):
        if coln == 1:
            ws.cell(row = rcounter, column = coln).value = day[0]
        elif coln == 2:
            ws.cell(row = rcounter, column = coln).value = capitaD[capitaIndex]
            capitaIndex = capitaIndex + 1
    rcounter = rcounter + 1


#biweekly percapita
rcounter = 2
ws.cell(row = 1, column = 4).value = 'biweekly-timestamp'
ws.cell(row = 1, column = 5).value = 'biweekly-percapita (per 100,000)'

for date in capitaBW:
    for coln in range(4,6):
        if coln == 4:
            ws.cell(row = rcounter, column = coln).value = date
        elif coln == 5:
            ws.cell(row = rcounter, column = coln).value = capitaBW[date]
    rcounter = rcounter + 1


wb.save(state + '-covid-per-capita.xlsx')
